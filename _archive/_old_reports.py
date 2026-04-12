import logging
from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from utils import calculate_total, format_ru_date, slugify_company

logger = logging.getLogger(__name__)

HEADERS = [
    "№",
    "Номер контейнера",
    "Тип",
    "Дата прибытия",
    "Дата вывоза",
    "Дней хранения",
    "Сумма $",
]

# Ширины столбцов (символы)
COL_WIDTHS = [5, 18, 8, 16, 16, 16, 12]

# Цвет заголовка
HEADER_FILL = PatternFill(start_color="2F75B6", end_color="2F75B6", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TOTAL_FONT = Font(bold=True, size=11)


def build_report(
    company_row,
    containers: list,
    report_kind: str,
    out_dir: Path,
) -> Path:
    """
    Генерирует xlsx-отчёт.

    company_row — Row с полями: name, entry_fee, free_days, storage_rate, storage_period_days
    containers  — список Row из containers_by_month
    report_kind — 'all' | 'departed'
    out_dir     — директория для сохранения файла
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Отчёт"

    # Заголовок
    ws.append(HEADERS)
    for col_idx, cell in enumerate(ws[1], 1):
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[cell.column_letter].width = COL_WIDTHS[col_idx - 1]
    ws.row_dimensions[1].height = 20

    total_sum = 0.0
    for idx, c in enumerate(containers, 1):
        arrival = date.fromisoformat(c["arrival_date"])
        departure = date.fromisoformat(c["departure_date"]) if c["departure_date"] else None
        days, total = calculate_total(
            arrival,
            departure,
            company_row["entry_fee"],
            company_row["free_days"],
            company_row["storage_rate"],
            company_row["storage_period_days"],
        )
        ws.append([
            idx,
            c["number"],
            c["type"],
            format_ru_date(arrival),
            format_ru_date(departure) if departure else "",
            days,
            total,
        ])
        # Выравнивание числовых ячеек
        row = ws[ws.max_row]
        for cell in row:
            cell.alignment = Alignment(horizontal="center")
        total_sum += total

    # Строка итогов с визуальным отступом в одну строку
    last_row = ws.max_row + 2
    ws.cell(row=last_row, column=6, value="Итого:")
    ws.cell(row=last_row, column=7, value=round(total_sum, 2))
    ws.cell(row=last_row, column=6).font = TOTAL_FONT
    ws.cell(row=last_row, column=7).font = TOTAL_FONT
    ws.cell(row=last_row, column=6).alignment = Alignment(horizontal="right")
    ws.cell(row=last_row, column=7).alignment = Alignment(horizontal="center")

    # Имя файла
    ts = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    filename = f"report_{slugify_company(company_row['name'])}_{report_kind}_{ts}.xlsx"
    path = out_dir / filename
    out_dir.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    logger.info("Отчёт сохранён: %s (%d строк, итого $%.2f)", path, len(containers), total_sum)
    return path

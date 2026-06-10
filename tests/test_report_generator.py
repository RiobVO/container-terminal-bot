"""Тесты генерации xlsx-отчёта: листы-месяцы, сводный лист, итоги."""
import pytest
from openpyxl import load_workbook

from services.report_generator import (
    EMPTY_MESSAGE,
    EMPTY_SHEET,
    HEADERS,
    build_report,
)

SETTINGS = {
    "default_entry_fee": 10.0,
    "default_free_days": 5,
    "default_storage_rate": 20.0,
    "default_storage_period_days": 30,
}


def _make_container(**over) -> dict:
    """Контейнер-словарь со всеми полями, нужными генератору и калькулятору."""
    base = {
        "number": "TEMU1234567",
        "display_number": "TEMU 1234567",
        "company_name": "Ромашка",
        "type": "40HQ",
        "status": "departed",
        "arrival_date": "2026-05-01 10:00:00",
        "departure_date": "2026-05-11 10:00:00",
        "comp_entry_fee": None,
        "comp_free_days": None,
        "comp_storage_rate": None,
        "comp_storage_period_days": None,
    }
    base.update(over)
    return base


class _RowLike:
    """Имитация aiosqlite.Row: только __getitem__, без .get()."""

    def __init__(self, data: dict):
        self._data = data

    def __getitem__(self, key):
        return self._data[key]


def test_empty_report_single_stub_sheet(tmp_path):
    """Пустой список — один лист «Нет данных» с сообщением."""
    path = build_report([], SETTINGS, tmp_path, "empty.xlsx")

    wb = load_workbook(path)
    assert wb.sheetnames == [EMPTY_SHEET]
    assert wb[EMPTY_SHEET].cell(row=1, column=1).value == EMPTY_MESSAGE


def test_invalid_group_field_raises(tmp_path):
    with pytest.raises(ValueError):
        build_report([], SETTINGS, tmp_path, "x.xlsx", group_field="status")


def test_months_summary_and_totals(tmp_path):
    """Группировка по месяцам прибытия, сводный лист, итоговая строка."""
    may = _make_container()  # days=10, billable=5, periods=1 → total 30
    june = _make_container(
        number="MSKU1234567",
        display_number="MSKU 1234567",
        company_name=None,
        arrival_date="2026-06-02",  # формат без времени
        departure_date="2026-06-05",
    )  # days=3, billable=0 → total 10
    # Непарсимая дата прибытия: не попадает в месячные листы,
    # но присутствует в сводном (in_transit — калькулятор вернёт нули)
    broken = _make_container(
        number="BROK1234567",
        display_number="BROK 1234567",
        status="in_transit",
        arrival_date="не дата",
        departure_date=None,
    )

    path = build_report(
        [june, may, broken], SETTINGS, tmp_path, "report.xlsx",
        summary_sheet_name="Сводный",
    )

    wb = load_workbook(path)
    assert wb.sheetnames == ["Сводный", "05.2026", "06.2026"]

    summary = wb["Сводный"]
    # Шапка + 3 строки данных + ИТОГО
    assert summary.max_row == 5
    assert [c.value for c in summary[1]] == list(HEADERS)
    # Сортировка по arrival_date asc; контейнер без даты — в конце
    assert summary.cell(row=2, column=2).value == "TEMU 1234567"
    assert summary.cell(row=3, column=2).value == "MSKU 1234567"
    assert summary.cell(row=4, column=2).value == "BROK 1234567"
    # Компания None → «—», статусы переведены
    assert summary.cell(row=3, column=3).value == "—"
    assert summary.cell(row=2, column=5).value == "Вывезен"
    assert summary.cell(row=4, column=5).value == "В пути"
    # Даты в формате дд.мм.гггг; пустая строка сохраняется openpyxl как None
    assert summary.cell(row=2, column=6).value == "01.05.2026"
    assert summary.cell(row=2, column=7).value == "11.05.2026"
    assert summary.cell(row=4, column=6).value is None
    # Итоговая строка
    assert summary.cell(row=5, column=2).value == "ИТОГО:"
    assert summary.cell(row=5, column=8).value == 13   # дни: 10 + 3 + 0
    assert summary.cell(row=5, column=9).value == 1    # периоды
    assert summary.cell(row=5, column=13).value == 20.0
    assert summary.cell(row=5, column=14).value == 40.0  # 30 + 10 + 0

    # Месячные листы: по одному контейнеру + шапка + итог
    for sheet, number, total in (
        ("05.2026", "TEMU 1234567", 30.0),
        ("06.2026", "MSKU 1234567", 10.0),
    ):
        ws = wb[sheet]
        assert ws.max_row == 3
        assert ws.cell(row=2, column=1).value == 1  # нумерация с 1
        assert ws.cell(row=2, column=2).value == number
        assert ws.cell(row=3, column=14).value == total


def test_group_by_departure_and_row_sorting(tmp_path):
    """Листы по departure_date, строки внутри — по arrival_date asc."""
    late_arrival = _make_container(
        number="BBBB1234567",
        display_number="BBBB 1234567",
        arrival_date="2026-05-20 10:00:00",
        departure_date="2026-06-03 10:00:00",
    )
    early_arrival = _make_container(
        number="AAAA1234567",
        display_number="AAAA 1234567",
        arrival_date="2026-05-02 10:00:00",
        departure_date="2026-06-25 10:00:00",
    )

    path = build_report(
        [late_arrival, early_arrival], SETTINGS, tmp_path, "departed.xlsx",
        group_field="departure_date",
    )

    wb = load_workbook(path)
    # Оба вывезены в июне — один лист, без сводного
    assert wb.sheetnames == ["06.2026"]
    ws = wb["06.2026"]
    assert ws.cell(row=2, column=2).value == "AAAA 1234567"
    assert ws.cell(row=3, column=2).value == "BBBB 1234567"


def test_row_without_get_method(tmp_path):
    """Строка без .get() (как aiosqlite.Row): доступ через [], пропуски → «—»."""
    row = _RowLike({
        # ключей type и company_name нет — _get вернёт None
        "display_number": "ROWW 1234567",
        "status": "on_terminal",
        "arrival_date": "2026-06-01 10:00:00",
        "departure_date": None,
    })

    path = build_report([row], SETTINGS, tmp_path, "rowlike.xlsx")

    wb = load_workbook(path)
    ws = wb["06.2026"]
    assert ws.cell(row=2, column=2).value == "ROWW 1234567"
    assert ws.cell(row=2, column=3).value == "—"  # company_name отсутствует
    assert ws.cell(row=2, column=4).value == "—"  # type отсутствует
    assert ws.cell(row=2, column=5).value == "На терминале"

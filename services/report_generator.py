"""Генерация xlsx-отчётов с разбивкой по листам-месяцам.

Один файл = несколько листов, по одному на каждый месяц. Поле, по которому
определяется «месяц контейнера», передаётся параметром ``group_field``:
``arrival_date`` — для отчётов по активным и по активным+вывезенным;
``departure_date`` — для отчёта по вывезенным. Если контейнеров нет —
создаётся один лист «Нет данных» (xlsx не может быть пустым).

Генератор ничего не знает о фильтрации: её обязан сделать вызывающий
хэндлер. На вход подаётся уже отфильтрованный список контейнеров, и
ожидается, что все они имеют значение в ``group_field``.
"""
import logging
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from services.calculator import calculate_container_cost

logger = logging.getLogger(__name__)

# Порядок колонок зафиксирован здесь: индексы ниже (MONEY_COLS, TOTAL_COLS,
# DATE_COLS) — 1-based и должны оставаться в синхроне с HEADERS.
HEADERS: tuple[str, ...] = (
    "№",
    "Номер контейнера",
    "Компания",
    "Тип",
    "Статус",
    "Дата прибытия",
    "Дата вывоза",
    "Дней на терминале",
    "Периодов к оплате",
    "Стоимость входа $",
    "Ставка хранения $",
    "Период начисления (дн.)",
    "Сумма хранения $",
    "К оплате $",
)

# 1-based индексы колонок для форматирования и агрегации
MONEY_COLS: tuple[int, ...] = (10, 11, 13, 14)
DATE_COLS: tuple[int, ...] = (6, 7)
# Итоги по колонкам: дни, периоды, сумма хранения, к оплате
TOTAL_COLS: tuple[int, ...] = (8, 9, 13, 14)

MONEY_FMT = '#,##0.00 "$"'
DATE_FMT = "DD.MM.YYYY"

STATUS_MAP = {
    "in_transit": "В пути",
    "on_terminal": "На терминале",
    "departed": "Вывезен",
}

EMPTY_SHEET = "Нет данных"
EMPTY_MESSAGE = "Нет контейнеров за выбранный период."

# Разрешённые значения group_field: через это имя достаётся месяц контейнера
# для группировки по листам и для сортировки строк внутри листа.
_GROUP_FIELD_ARRIVAL = "arrival_date"
_GROUP_FIELD_DEPARTURE = "departure_date"

# Стили
_HEADER_FILL = PatternFill(
    start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"
)
_HEADER_FONT = Font(bold=True, size=11)
_TOTAL_FILL = PatternFill(
    start_color="BFBFBF", end_color="BFBFBF", fill_type="solid"
)
_TOTAL_FONT = Font(bold=True, size=11)
_thin = Side(style="thin", color="999999")
_BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
_CENTER = Alignment(horizontal="center", vertical="center")

_MIN_COL_WIDTH = 10
_MAX_COL_WIDTH = 30


def _get(row, key):
    """Универсальный доступ к полю строки БД или dict-подобного объекта."""
    if hasattr(row, "get"):
        return row.get(key)
    try:
        return row[key]
    except (KeyError, IndexError):
        return None


def _parse_date(val: str | None) -> datetime | None:
    """Парсит дату из строки БД. Поддерживает форматы с временем и без."""
    if not val:
        return None
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try:
            return datetime.strptime(val, fmt)
        except ValueError:
            continue
    return None


def _sheet_key_for_container(c, group_field: str) -> str | None:
    """«MM.YYYY» по значению ``group_field`` или None, если даты нет.

    Контейнеры без значения в group_field не группируются и не попадают
    в отчёт: вызывающий хэндлер обязан отфильтровать такие строки.
    """
    dt = _parse_date(_get(c, group_field))
    if dt is None:
        return None
    return f"{dt.month:02d}.{dt.year}"


def _group_by_month(
    containers: list, group_field: str
) -> dict[str, list]:
    """Группирует контейнеры по месяцу от ``group_field``.

    Листы упорядочены по возрастанию даты (самый старый — первый).
    Контейнеры без даты в ``group_field`` молча отбрасываются: в новой
    спеке ни один из трёх типов отчёта не содержит такие записи.
    """
    buckets: dict[str, list] = {}
    for c in containers:
        key = _sheet_key_for_container(c, group_field)
        if key is None:
            continue
        buckets.setdefault(key, []).append(c)

    def _month_sort_key(name: str) -> tuple[int, int]:
        mm, yyyy = name.split(".")
        return (int(yyyy), int(mm))

    return {k: buckets[k] for k in sorted(buckets, key=_month_sort_key)}


def _sort_rows(containers: list) -> list:
    """Строки внутри листа: по дате прибытия asc, затем компания, номер.

    Сортировка всегда идёт по ``arrival_date`` независимо от того, по
    какому полю группируется лист. В отчёте «🔴 Только вывезенные» листы
    группируются по ``departure_date``, но строки внутри листа всё равно
    должны идти от самого старого прибытия к самому свежему — это явное
    требование спецификации.
    """
    def key(c):
        dt = _parse_date(_get(c, "arrival_date")) or datetime.max
        company = (_get(c, "company_name") or "").lower()
        number = (_get(c, "display_number") or "").upper()
        return (dt, company, number)
    return sorted(containers, key=key)


def _autosize_columns(ws: Worksheet, n_cols: int) -> None:
    """Подгоняет ширину колонок по содержимому в пределах [10; 30]."""
    for col_idx in range(1, n_cols + 1):
        letter = get_column_letter(col_idx)
        max_len = 0
        for cell in ws[letter]:
            val = cell.value
            if val is None:
                continue
            text = (
                val.strftime("%d.%m.%Y")
                if isinstance(val, datetime)
                else str(val)
            )
            if len(text) > max_len:
                max_len = len(text)
        width = max(_MIN_COL_WIDTH, min(_MAX_COL_WIDTH, max_len + 2))
        ws.column_dimensions[letter].width = width


def _write_header(ws: Worksheet) -> None:
    """Шапка листа с заливкой, жирным шрифтом, границами и freeze panes."""
    ws.append(list(HEADERS))
    for cell in ws[1]:
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _CENTER
        cell.border = _BORDER
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"


def _write_data_row(
    ws: Worksheet, idx: int, c, settings: dict
) -> dict:
    """Пишет одну строку данных. Возвращает cost-словарь для агрегации."""
    cost = calculate_container_cost(
        c, settings,
        comp_entry_fee=_get(c, "comp_entry_fee"),
        comp_free_days=_get(c, "comp_free_days"),
        comp_storage_rate=_get(c, "comp_storage_rate"),
        comp_storage_period_days=_get(c, "comp_storage_period_days"),
    )

    company_name = _get(c, "company_name") or "—"
    status_human = STATUS_MAP.get(c["status"], c["status"])

    arrival_cell = _parse_date(_get(c, "arrival_date")) or ""
    departure_cell = _parse_date(_get(c, "departure_date")) or ""

    ws.append([
        idx,
        _get(c, "display_number"),
        company_name,
        _get(c, "type") or "—",
        status_human,
        arrival_cell,
        departure_cell,
        cost["days"],
        cost["periods"],
        cost["entry_fee"],
        cost["storage_rate"],
        cost["period_days"],
        cost["storage"],
        cost["total"],
    ])

    row = ws[ws.max_row]
    for cell in row:
        cell.alignment = _CENTER
    for col in MONEY_COLS:
        row[col - 1].number_format = MONEY_FMT
    for col in DATE_COLS:
        row[col - 1].number_format = DATE_FMT

    return cost


def _write_totals(ws: Worksheet, running: dict[int, float]) -> None:
    """Итоговая строка в конце листа: «ИТОГО:» + суммы по TOTAL_COLS."""
    last_row = ws.max_row + 1
    ws.cell(row=last_row, column=2, value="ИТОГО:")
    for col in TOTAL_COLS:
        ws.cell(row=last_row, column=col, value=running.get(col, 0))

    for col_idx in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=last_row, column=col_idx)
        cell.font = _TOTAL_FONT
        cell.fill = _TOTAL_FILL
        cell.alignment = _CENTER

    for col in MONEY_COLS:
        if col in TOTAL_COLS:
            ws.cell(row=last_row, column=col).number_format = MONEY_FMT


def _fill_sheet(
    ws: Worksheet,
    containers: list,
    settings: dict,
) -> None:
    """Заполняет один лист: шапка, строки (с нумерацией с 1), итог."""
    _write_header(ws)

    # Агрегаты: дни и периоды — целые, денежные — float с округлением в конце.
    running: dict[int, float] = {8: 0, 9: 0, 13: 0.0, 14: 0.0}

    for idx, c in enumerate(_sort_rows(containers), 1):
        cost = _write_data_row(ws, idx, c, settings)
        running[8] += int(cost["days"] or 0)
        running[9] += int(cost["periods"] or 0)
        running[13] += float(cost["storage"] or 0.0)
        running[14] += float(cost["total"] or 0.0)

    running[13] = round(running[13], 2)
    running[14] = round(running[14], 2)

    if ws.max_row > 1:
        _write_totals(ws, running)

    _autosize_columns(ws, len(HEADERS))


def build_report(
    containers: list,
    settings: dict[str, float],
    out_dir: Path,
    filename: str,
    *,
    group_field: str = _GROUP_FIELD_ARRIVAL,
    summary_sheet_name: str | None = None,
) -> Path:
    """Генерирует xlsx-отчёт с разбивкой по листам-месяцам.

    Контейнеры уже отфильтрованы вызывающим кодом (по компании/статусу).
    ``group_field`` — имя поля БД, по которому берётся месяц листа:
    ``arrival_date`` для активных/микс-отчётов, ``departure_date`` —
    для отчёта по вывезенным.

    ``summary_sheet_name`` — если задан и контейнеры не пусты, перед
    месячными листами вставляется сводный лист с этим именем, куда
    попадают все контейнеры одним списком с общей итоговой строкой.
    Сортировка и формат — как у обычных листов. Осознанное дублирование:
    каждый контейнер присутствует и на сводном, и на своём месячном.

    Если итоговый список пуст — создаётся единственный лист «Нет данных».
    """
    if group_field not in (_GROUP_FIELD_ARRIVAL, _GROUP_FIELD_DEPARTURE):
        raise ValueError(f"Недопустимое group_field: {group_field!r}")

    wb = Workbook()
    # Удаляем дефолтный лист — дальше создаём свои с явными именами.
    wb.remove(wb.active)

    groups = _group_by_month(containers, group_field) if containers else {}

    if not groups:
        ws = wb.create_sheet(title=EMPTY_SHEET)
        ws.cell(row=1, column=1, value=EMPTY_MESSAGE)
        ws.column_dimensions["A"].width = len(EMPTY_MESSAGE) + 2
    else:
        if summary_sheet_name:
            ws = wb.create_sheet(title=summary_sheet_name)
            _fill_sheet(ws, containers, settings)
        for sheet_name, rows in groups.items():
            ws = wb.create_sheet(title=sheet_name)
            _fill_sheet(ws, rows, settings)

    out_dir.mkdir(parents=True, exist_ok=True)
    path = out_dir / filename
    wb.save(path)
    logger.info(
        "Отчёт: %s (контейнеров: %d, листов: %d, group=%s, summary=%s)",
        path, len(containers), len(wb.sheetnames), group_field,
        summary_sheet_name or "—",
    )
    return path

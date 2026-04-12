"""Smoke-тест раздела отчётов: прогоняет все 6 комбинаций через реальный
путь `db.containers.fetch_for_report` → `services.report_generator.build_report`
и валидирует структуру сгенерированных xlsx-файлов.

Что проверяется по каждому файлу:
- 14 колонок, нет колонки «Дата регистрации»;
- freeze_panes = A2;
- нумерация в колонке № сквозная 1..N внутри каждого листа;
- итоговая строка «ИТОГО:» есть и её суммы по колонкам 8/9/13/14
  совпадают с суммой строк того же листа (±0.01);
- сортировка листов по возрастанию месяца;
- для типа «🟢 Активные» первым листом идёт «Все активные», для двух
  остальных типов такого листа нет;
- кросс-месячный контейнер (если такой есть в базе): отсутствует в
  «active», в «mixed» — на листе месяца прибытия со статусом «Вывезен»,
  в «departed» — на листе месяца вывоза.

Запуск:
    python -m scripts.smoke_reports
    python -m scripts.smoke_reports --keep     # не удалять файлы после проверки
    python -m scripts.smoke_reports --db other.db --out-dir smoke_out

Коды выхода: 0 — всё OK, 1 — были ошибки. Полный лог инвариантов идёт
в stdout с префиксами [OK] / [FAIL].
"""
from __future__ import annotations

import argparse
import asyncio
import logging
import sys
from datetime import datetime
from pathlib import Path

# На Windows консоль по умолчанию cp1251 — ломается на кириллице и юникод-
# символах типа стрелок. Принудительно переводим stdout/stderr в UTF-8.
for _stream in (sys.stdout, sys.stderr):
    if hasattr(_stream, "reconfigure"):
        try:
            _stream.reconfigure(encoding="utf-8")
        except Exception:
            pass

from openpyxl import load_workbook

# Корень проекта должен быть в sys.path, чтобы импортировались пакеты.
_PROJECT_ROOT = Path(__file__).resolve().parent.parent
if str(_PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(_PROJECT_ROOT))

import db  # noqa: E402
from config import load_config  # noqa: E402
from db import containers as db_cont  # noqa: E402
from db.settings import get_all_settings  # noqa: E402
from handlers.reports import _REPORT_SPECS, _slugify  # noqa: E402
from services.report_generator import (  # noqa: E402
    EMPTY_SHEET,
    HEADERS,
    build_report,
)

logging.basicConfig(level=logging.WARNING, format="%(levelname)s %(name)s: %(message)s")

# Для записи результатов
_PASS: list[str] = []
_FAIL: list[str] = []


def _ok(msg: str) -> None:
    _PASS.append(msg)
    print(f"[OK]   {msg}")


def _fail(msg: str) -> None:
    _FAIL.append(msg)
    print(f"[FAIL] {msg}")


def _parse_date(val) -> datetime | None:
    if val is None or val == "":
        return None
    if isinstance(val, datetime):
        return val
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try:
            return datetime.strptime(val, fmt)
        except ValueError:
            continue
    return None


def _month_key(dt: datetime) -> str:
    return f"{dt.month:02d}.{dt.year}"


def _expected_month_sheets(
    containers: list, group_field: str
) -> list[str]:
    """Ожидаемый список месячных листов (без сводного), отсортированный."""
    months = set()
    for c in containers:
        dt = _parse_date(c[group_field])
        if dt is not None:
            months.add(_month_key(dt))

    def sort_key(name: str) -> tuple[int, int]:
        mm, yyyy = name.split(".")
        return (int(yyyy), int(mm))

    return sorted(months, key=sort_key)


def _verify_sheet(ws, label: str) -> tuple[int, dict[int, float]]:
    """Проверяет шапку, нумерацию, freeze panes, итоговую строку.

    Возвращает (кол-во строк данных, sums по колонкам 8/9/13/14),
    чтобы вызывающий мог сверить это с данными от БД.
    """
    # Шапка
    headers = [
        ws.cell(row=1, column=i).value for i in range(1, ws.max_column + 1)
    ]
    if len(headers) != len(HEADERS):
        _fail(f"{label}: колонок {len(headers)}, ожидалось {len(HEADERS)}")
    if "Дата регистрации" in headers:
        _fail(f"{label}: в шапке присутствует «Дата регистрации»")
    if ws.freeze_panes != "A2":
        _fail(f"{label}: freeze_panes={ws.freeze_panes!r}, ожидалось 'A2'")

    # Строки данных и итоговая строка
    data_rows: list[int] = []
    sums = {8: 0.0, 9: 0.0, 13: 0.0, 14: 0.0}
    total_row_idx: int | None = None

    for r in range(2, ws.max_row + 1):
        col2 = ws.cell(row=r, column=2).value
        if col2 == "ИТОГО:":
            total_row_idx = r
            continue
        n = ws.cell(row=r, column=1).value
        if n is None:
            continue
        data_rows.append(int(n))
        for col in sums:
            v = ws.cell(row=r, column=col).value
            if v is not None:
                sums[col] += float(v)

    # Нумерация 1..N
    expected_numbers = list(range(1, len(data_rows) + 1))
    if data_rows != expected_numbers:
        _fail(
            f"{label}: нумерация {data_rows[:10]}... != 1..{len(data_rows)}"
        )

    # Итоговая строка должна быть, если есть данные
    if data_rows and total_row_idx is None:
        _fail(f"{label}: нет строки «ИТОГО:» при {len(data_rows)} данных")
    elif data_rows and total_row_idx is not None:
        for col in (8, 9, 13, 14):
            cell_val = ws.cell(row=total_row_idx, column=col).value
            actual = float(cell_val or 0)
            expected = round(sums[col], 2)
            if abs(actual - expected) > 0.01:
                _fail(
                    f"{label}: ИТОГО col{col}={actual}, сумма строк={expected}"
                )

    return len(data_rows), sums


def _verify_file(
    path: Path,
    containers: list,
    group_field: str,
    summary_sheet_name: str | None,
    scenario_label: str,
) -> None:
    """Полная валидация одного сгенерированного xlsx."""
    wb = load_workbook(path)
    sheets = wb.sheetnames

    expected_monthly = _expected_month_sheets(containers, group_field)

    if not containers:
        if sheets == [EMPTY_SHEET]:
            _ok(f"{scenario_label}: пустой отчёт → лист «{EMPTY_SHEET}»")
        else:
            _fail(f"{scenario_label}: ожидался один лист «{EMPTY_SHEET}», "
                  f"получены {sheets}")
        return

    # Ожидаемый порядок листов
    expected_sheets = list(expected_monthly)
    if summary_sheet_name:
        expected_sheets = [summary_sheet_name] + expected_sheets

    if sheets != expected_sheets:
        _fail(
            f"{scenario_label}: порядок листов {sheets} != {expected_sheets}"
        )
    else:
        _ok(f"{scenario_label}: порядок листов {sheets}")

    # Проверка сводного листа: количество строк = len(containers),
    # суммы дней/периодов/хранения/total должны совпадать с суммой по всем
    # месячным листам.
    if summary_sheet_name and summary_sheet_name in wb.sheetnames:
        sum_rows, sum_totals = _verify_sheet(
            wb[summary_sheet_name], f"{scenario_label}/{summary_sheet_name}"
        )
        if sum_rows != len(containers):
            _fail(
                f"{scenario_label}/{summary_sheet_name}: строк {sum_rows}, "
                f"ожидалось {len(containers)}"
            )
        else:
            _ok(
                f"{scenario_label}/{summary_sheet_name}: "
                f"{sum_rows} строк = всего активных"
            )

        # Суммы по сводному = суммы по всем месячным
        monthly_totals = {8: 0.0, 9: 0.0, 13: 0.0, 14: 0.0}
        for name in expected_monthly:
            _, totals = _verify_sheet(
                wb[name], f"{scenario_label}/{name}"
            )
            for k in monthly_totals:
                monthly_totals[k] += totals[k]
        for col in (8, 9, 13, 14):
            a = round(sum_totals[col], 2)
            b = round(monthly_totals[col], 2)
            if abs(a - b) > 0.01:
                _fail(
                    f"{scenario_label}: сводный col{col}={a} != "
                    f"сумма месячных={b}"
                )
        # Отдельно отметим, что баланс сведён
        _ok(f"{scenario_label}: суммы сводного = сумма месячных")
    else:
        # Без сводного — валидируем каждый месячный лист по одному разу
        for name in expected_monthly:
            _verify_sheet(wb[name], f"{scenario_label}/{name}")


async def _run_scenario(
    report_type: str,
    company,
    out_dir: Path,
    settings: dict,
) -> Path:
    """Сбор данных + запуск build_report. Возвращает путь к файлу."""
    spec = _REPORT_SPECS[report_type]
    statuses = spec["statuses"]
    group_field = spec["group_field"]
    summary_sheet = spec["summary_sheet"]

    company_id = company["id"] if company else None
    company_name = company["name"] if company else None

    containers = await db_cont.fetch_for_report(
        statuses=statuses, company_id=company_id
    )

    ts = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    if company_name is None:
        filename = f"{spec['file_prefix_all']}_{ts}_SMOKE.xlsx"
    else:
        filename = (
            f"{spec['file_prefix_company']}_{_slugify(company_name)}"
            f"_{ts}_SMOKE.xlsx"
        )

    path = build_report(
        list(containers),
        settings,
        out_dir,
        filename,
        group_field=group_field,
        summary_sheet_name=summary_sheet,
    )

    label = (
        f"{report_type}/all" if company is None
        else f"{report_type}/company={company_name}"
    )
    _verify_file(path, list(containers), group_field, summary_sheet, label)
    return path


async def _pick_test_company(report_type: str):
    """Возвращает первую компанию, у которой есть контейнеры нужных статусов."""
    spec = _REPORT_SPECS[report_type]
    async with db.get_db() as conn:
        import aiosqlite
        conn.row_factory = aiosqlite.Row
        placeholders = ",".join("?" * len(spec["statuses"]))
        row = await (
            await conn.execute(
                f"SELECT co.id, co.name FROM companies co "
                f"JOIN containers c ON c.company_id=co.id "
                f"WHERE c.status IN ({placeholders}) "
                f"GROUP BY co.id, co.name "
                f"ORDER BY co.name COLLATE NOCASE LIMIT 1",
                list(spec["statuses"]),
            )
        ).fetchone()
    return dict(row) if row else None


async def _cross_month_check(out_dir: Path, settings: dict) -> None:
    """Если в БД есть контейнер, у которого месяц прибытия != месяц вывоза,
    проверяем его расположение во всех трёх типах отчётов."""
    async with db.get_db() as conn:
        import aiosqlite
        conn.row_factory = aiosqlite.Row
        rows = await (
            await conn.execute(
                "SELECT display_number, company_id, arrival_date, departure_date "
                "FROM containers WHERE status='departed' "
                "AND arrival_date IS NOT NULL AND departure_date IS NOT NULL"
            )
        ).fetchall()

    target = None
    for r in rows:
        a = _parse_date(r["arrival_date"])
        d = _parse_date(r["departure_date"])
        if a and d and (a.year, a.month) != (d.year, d.month):
            target = dict(r)
            target["_arr"] = a
            target["_dep"] = d
            break

    if target is None:
        print("[SKIP] кросс-месячного контейнера не найдено, проверку пропускаем")
        return

    num = target["display_number"]
    arr_sheet = _month_key(target["_arr"])
    dep_sheet = _month_key(target["_dep"])
    print(
        f"\n=== кросс-месячный контейнер: {num} "
        f"(arr={arr_sheet}, dep={dep_sheet}) ==="
    )

    def find_in_wb(path: Path, number: str) -> list[tuple[str, str]]:
        """Возвращает список (лист, статус) с этим номером."""
        wb = load_workbook(path)
        found = []
        for name in wb.sheetnames:
            if name == EMPTY_SHEET:
                continue
            ws = wb[name]
            for r in range(2, ws.max_row + 1):
                if ws.cell(row=r, column=2).value == number:
                    status = ws.cell(row=r, column=5).value
                    found.append((name, status))
                    break
        return found

    # Построим 3 отчёта по всем компаниям и поищем контейнер.
    for rt in ("active", "mixed", "departed"):
        path = await _run_scenario(rt, None, out_dir, settings)
        hits = find_in_wb(path, num)
        if rt == "active":
            if hits:
                _fail(f"{num}: присутствует в active ({hits}), не должен")
            else:
                _ok(f"{num}: отсутствует в active")
        elif rt == "mixed":
            expected = [(arr_sheet, "Вывезен")]
            if hits == expected:
                _ok(f"{num}: mixed → {hits[0]}")
            else:
                _fail(f"{num}: mixed → {hits}, ожидалось {expected}")
        elif rt == "departed":
            # В departed может быть на сводном (если он там есть), но у
            # departed summary_sheet=None, так что только месячный лист.
            expected = [(dep_sheet, "Вывезен")]
            if hits == expected:
                _ok(f"{num}: departed → {hits[0]}")
            else:
                _fail(f"{num}: departed → {hits}, ожидалось {expected}")


async def main(args: argparse.Namespace) -> int:
    cfg = load_config()
    db._DB_PATH = args.db or cfg.db_path
    print(f"Используется БД: {db._DB_PATH}")

    out_dir: Path = args.out_dir
    out_dir.mkdir(parents=True, exist_ok=True)

    settings = await get_all_settings()
    if not settings:
        print(
            "⚠️ В БД нет global_settings. Запусти бота хотя бы раз или "
            "добавь дефолты вручную."
        )
        return 1

    generated: list[Path] = []

    # Все 6 базовых комбинаций: тип × scope
    for report_type in ("active", "mixed", "departed"):
        # Режим «по всем»
        path = await _run_scenario(report_type, None, out_dir, settings)
        generated.append(path)

        # Режим «по одной компании» — берём подходящую
        company = await _pick_test_company(report_type)
        if company is None:
            print(f"[SKIP] {report_type}: нет компаний с такими статусами")
            continue
        path = await _run_scenario(report_type, company, out_dir, settings)
        generated.append(path)

    # Дополнительно — проверка кросс-месячного контейнера (если есть)
    await _cross_month_check(out_dir, settings)

    # Итог
    print()
    print("=" * 60)
    print(f"Итого: {len(_PASS)} OK, {len(_FAIL)} FAIL")
    print("=" * 60)

    if not args.keep:
        for p in generated:
            p.unlink(missing_ok=True)
        # Удалим только что сгенерированные для кросс-месячной проверки тоже
        for p in out_dir.glob("*_SMOKE.xlsx"):
            p.unlink(missing_ok=True)
        print("Файлы удалены (передай --keep чтобы оставить).")
    else:
        print(f"Файлы сохранены в {out_dir}/")

    return 0 if not _FAIL else 1


def _parse_args() -> argparse.Namespace:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--db", help="Путь к БД (по умолчанию — из .env)")
    ap.add_argument(
        "--out-dir", type=Path, default=Path("test_reports"),
        help="Куда класть сгенерированные файлы"
    )
    ap.add_argument(
        "--keep", action="store_true",
        help="Не удалять файлы после проверки",
    )
    return ap.parse_args()


if __name__ == "__main__":
    args = _parse_args()
    sys.exit(asyncio.run(main(args)))
